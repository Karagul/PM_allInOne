from openpyxl import load_workbook
import csv


class exchange_notice():


    def __init__(self, script_dir):


        self.en_matrix = script_dir + '\\EN_Templates\\EN_matrix.txt'
        self.en_template = script_dir + '\\EN_Templates\\EN_template.xlsx'
        # self.en_template = r'C:\Users\mavkri\Documents\PM_pyscripts\EN_Templates\EN_template.xlsx'
        self.getMatrix()


    def getData(self, csvFile):

        with open(csvFile, newline='') as f:
            reader = csv.reader(f, delimiter='|',
                                   escapechar='',
                                   quotechar="'")
            n = 0
            # for row in reader:
            #     if n == 1:
            #         csvData = row[:]
            #     n += 1
            for row in reader:
                if n == 1:
                    csvData = row
                elif n > 1:
                    if row:
                        csvData[-1] += ' ' + row[0].replace("'","")
                        csvData += row[1:]
                n += 1

        return csvData

    def getMatrix(self):
        
        with open(self.en_matrix, 'r') as f:
            self.csvMap = f.readlines()
        self.csvMap = [x.strip().split('\t') for x in self.csvMap]
        self.csvMap = [(x[0], x[1].split(';')) for x in self.csvMap]
        # print (self.csvMap)

    def openWorkbook(self):
        wb = load_workbook(self.en_template)
        return wb


    def fillTemplate(self, workbook, csvData, col):
        
        # wb = load_workbook(self.en_template)
        ws = workbook['Sheet1']

        # dates = [20, 50, 51, 52, 53, 54, 55, 57]
        dates = [2,7]
        # 56 only year

        self.listingDate = str(csvData[58])

        for elem in self.csvMap:
            if (elem[1][0] != 'None'):
                # if elem[0] == '56':
                #     ws.cell(row=int(elem[0]), column=col).value = ((csvData[int(elem[1][0])])[:4]).strip()
                if int(elem[0]) in dates :
                    d = (csvData[int(elem[1][0])]).strip()
                    if str(d) != '0':                    
                        ws.cell(row=int(elem[0]), column=col).value = '-'.join([d[:4], d[4:6], d[6:]])
                    elif int(d) == 0:
                        ws.cell(row=int(elem[0]), column=col).value = '0'
                # elif len(elem[1]) > 1:
                #     tempText = ''
                #     for i in range(int(elem[1][0]), int(elem[1][1])):
                #         tempText += (str(csvData[i]) + ' ')
                #     ws.cell(row=int(elem[0]), column=col).value = tempText.strip()
                else:
                    ws.cell(row=int(elem[0]), column=col).value = (csvData[int(elem[1][0])]).strip()

        # extraTags = ['Termin / Term (number)',
        #              'Første emissionsdato / First subscription date',
        #              'Sidste emissionsdato / Last redemption date',
        #              'Trækningsberegningsdato / Date for calculation of drawings',
        #              'Trækningsdato / Date for drawing',
        #              'Publiceringsdato / Publication date',
        #              'Terminsdato / Payment date']
        # n = 88
        # x = 1
        # while True:
        #     if csvData[n] != '':
        #         ws.cell(row=(n+x), column=col).value = str(x)
        #         ws.cell(row=(n+x), column=1).value = extraTags[0]
        #         for i in range(6):
        #             # print (i)
        #             d = csvData[n+i]
        #             ws.cell(row=(n+x+1+i), column=col).value = '-'.join([d[:4], d[4:6], d[6:]])
        #             ws.cell(row=(n+x+1+i), column=1).value = extraTags[i+1]
        #         n += 6
        #         x += 1
        #     else:
        #         break

        # return wb


    def saveWorkbook(self,workbook, dest):
        
        workbook.save(dest + self.listingDate + '_KFD_Exchange_Notice.xlsx')