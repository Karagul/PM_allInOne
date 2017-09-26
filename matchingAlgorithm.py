from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import PyPDF2
import createFolders
import calculateDays
import edit_docx_file_test as edit_docx_file
from datetime import datetime
import win32com.client as win32
import os
import shutil



class matchingAlgorithm():

    def __init__(self, xlsxLocation='Evening Strikes.xlsx'):

        wb = load_workbook(filename=xlsxLocation, read_only=True)
        series = wb['Series']
        underlying = wb['Underlying']

        instrumentTypes = []
        for row in underlying.iter_rows(min_row=3, max_col=2):
            if row[1].value != None:
                instrumentTypes.append((row[0].value, row[1].value))

        generatedSeries = []
        for row in series.iter_rows(min_row=1, max_col=5):
            if (row[0].value and row[4].value) != None:
                generatedSeries.append((row[0].value, row[1].value, 
                                        row[2].value, row[3].value, 
                                        row[4].value))

        self.generatedAmount = series['H1'].value
        self.enNumber = str(series['H2'].value)
        self.enYear = str(series['H3'].value)[-2:]
        wb._archive.close()
        allLists = self.comparison(generatedSeries, instrumentTypes)
        self.matchedAmount = len(allLists[0])
        self.cleanup_matched(allLists[0])
        if self.generatedAmount != self.matchedAmount:
            self.cleanup_unmatched(allLists[1], allLists[2])
        else:
            self.create_docx_file()

    # function to check where two values
    # alphabeticaly stand with respect to each other
    def alphaCompare(self, value1, value2):
        # if they are the same return 'equal'
        if value1 == value2:
            return 'equal'
        # if value1 is above value2, return 'value1'
        elif sorted((value1, value2))[0] == value1:
            return 'value1'
        # if value2 is above value1, return 'value2'
        else:
            return 'value2'

    # function to compare two sorted lists
    def comparison(self, shortList, longList):

        # need only strikes from these markets
        self.markets = ['DAI', 'DAS',
                        'EUI',
                        'HXS', 'HXSOR',
                        'ICI', 'ICS',
                        'NNOI', 'NNOS',
                        'SES', 'SEI',
                        'USI', 'USS']
        # sort shortList
        # print (shortList)
        shortList = sorted(shortList, key=lambda x: x[4])
        # sort longList
        longList = sorted(longList)
        # array to contain all matches
        matchedList = []
        # array to contain left overs
        unknownList = []
        # unimportant list
        dismissList = []
        # define i and j for iteration algorithm
        i = 0
        j = 0
        # start cycle until one of the lists are checked
        while ((i < len(shortList)) and (j < len(longList))):

            # get the status of how these values look
            # with respect to each other
            status = self.alphaCompare(shortList[i][4], longList[j][0])

            # if the values are equal, append match array
            if status == 'equal':
                if longList[j][1] in self.markets:
                    matchedList.append((shortList[i][0], 
                                       shortList[i][1],
                                       shortList[i][2],
                                       shortList[i][3],
                                       longList[j][1]))
                else:
                    dismissList.append((shortList[i][0], 
                                       shortList[i][1],
                                       shortList[i][2],
                                       shortList[i][3]))
                i += 1
                # short list might contain multiples of the same value
                # we want to capture those as well
                # change longList index only when unique shortList value appears
                try:
                    if shortList[i][4] != shortList[i-1][4]:
                        j += 1
                except IndexError:
                    break
            # if shortList element is above longList elemnt alphabeticaly
            # we skip to the next value
            elif status == 'value1':
                unknownList.append((shortList[i][0], 
                                     shortList[i][1],
                                     shortList[i][2],
                                     shortList[i][3]))
                i += 1
            # reverse of the above
            else:
                j += 1

        # sort matched list
        matchedList = sorted(matchedList, key=lambda x: (x[4], x[0]))

        # return matched and unmatched elements
        return matchedList, dismissList, unknownList

    def cleanup_matched(self, matchedList):

        self.marketsMap = {0 : 'Danish Underlying',
                           1 : 'VINX30',
                           2 : 'Finnish Underlying',
                           3 : 'Icelandic Underlying',
                           4 : 'Norwegian Underlying',
                           5 : 'Swedish Underlying',
                           6 : 'Russian Underlying'}

        defaultLayout = ['Instrument Series', 'ISIN Code', 
                         'Expiration Date', 'Strike Price']

        self.sortedList = [[] for x in range(7)]
        for row in matchedList:
            if row[4] in self.markets[0:2]:
                self.sortedList[0].append(row[:4])
            elif row[4] in self.markets[2:3]:
                self.sortedList[1].append(row[:4])
            elif row[4] in self.markets[3:5]:
                self.sortedList[2].append(row[:4])
            elif row[4] in self.markets[5:7]:
                self.sortedList[3].append(row[:4])
            elif row[4] in self.markets[7:9]:
                self.sortedList[4].append(row[:4])
            elif row[4] in self.markets[9:11]:
                self.sortedList[5].append(row[:4])
            elif row[4] in self.markets[11:13]:
                self.sortedList[6].append(row[:4])
        # create new excel file
        wb = Workbook()

        foldersObj = createFolders.config_folders(*calculateDays.
                                                  next_business_day().
                                                  get_day_month())

        self.monthFolder = foldersObj.create_folders()
        
        self.day = foldersObj.day
        self.today = foldersObj.today

        self.fileName = 'New_Strikes_' + self.day + '.xlsx'

        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.column_dimensions['A'].width = 21
        ws1.column_dimensions['B'].width = 14
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12

        ft = Font(color='0000FF')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        n = 1
        for num, oneList in enumerate(self.sortedList):
            if oneList:
                ws1.cell(row=n, column=1).value = self.marketsMap[num]
                ws1.cell(row=n, column=1).font = ft
                ws1.cell(row=n, column=1).border = thin_border
                n += 1
                k = 1
                for value in defaultLayout:
                    ws1.cell(row=n, column=k).value = value
                    ws1.cell(row=n, column=k).border = thin_border
                    k += 1
                n += 1
                for singleLine in oneList:
                    k = 1
                    for value in singleLine:
                        if k == 3:
                            ws1.cell(row=n, column=k).value = value
                            ws1.cell(row=n, column=k).number_format = 'YYYY-MM-DD'
                            ws1.cell(row=n, column=k).border = thin_border
                        else:
                            ws1.cell(row=n, column=k).value = value
                            ws1.cell(row=n, column=k).border = thin_border
                        k += 1
                    n += 1

        wb.save(self.monthFolder + self.fileName)

    def cleanup_unmatched(self, dismissList, unknownList):

        defaultLayout = ['Instrument Series', 'ISIN Code', 
                         'Expiration Date', 'Strike Price']

        # create new excel file
        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 14
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 12

        ft = Font(color='0000FF')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        n = 1
        if unknownList:
            ws1.cell(row=n, column=1).value = 'Class and Type is missing'
            ws1.cell(row=n, column=1).font = ft
            ws1.cell(row=n, column=1).border = thin_border
            n += 1
            k = 1
            for value in defaultLayout:
                ws1.cell(row=n, column=k).value = value
                ws1.cell(row=n, column=k).border = thin_border
                k += 1
            n += 1
            for singleLine in unknownList:
                k = 1
                for value in singleLine:
                    if k == 3:
                        ws1.cell(row=n, column=k).value = value
                        ws1.cell(row=n, column=k).number_format = 'YYYY-MM-DD'
                        ws1.cell(row=n, column=k).border = thin_border
                    else:
                        ws1.cell(row=n, column=k).value = value
                        ws1.cell(row=n, column=k).border = thin_border
                    k += 1
                n += 1

        if dismissList:
            ws1.cell(row=n, column=1).value = 'Unnecessary strikes'
            ws1.cell(row=n, column=1).font = ft
            ws1.cell(row=n, column=1).border = thin_border
            n += 1
            k = 1
            for value in defaultLayout:
                ws1.cell(row=n, column=k).value = value
                ws1.cell(row=n, column=k).border = thin_border
                k += 1
            n += 1
            for singleLine in dismissList:
                k = 1
                for value in singleLine:
                    if k == 3:
                        ws1.cell(row=n, column=k).value = value
                        ws1.cell(row=n, column=k).number_format = 'YYYY-MM-DD'
                        ws1.cell(row=n, column=k).border = thin_border
                    else:
                        ws1.cell(row=n, column=k).value = value
                        ws1.cell(row=n, column=k).border = thin_border
                    k += 1
                n += 1

        wb.save(self.monthFolder + self.fileName[:-5] + '_Dismissed.xlsx')

    def create_docx_file(self):
        Super_Duper = os.path.expanduser('~\Documents\SUPER DUPER\\')
        if not os.path.exists(Super_Duper):
            os.makedirs(Super_Duper)

        scriptDir = os.path.dirname(__file__)

        defaultLayout = ['Instrument Series', 'ISIN Code', 
                         'Expiration Date', 'Strike Price']

        fileName = 'New_Strikes_Template.docx'
        absScriptPath = scriptDir + '\Templates\\' + fileName
        document = edit_docx_file.create_docx_file(absScriptPath)

        tomorrow = str(calculateDays.next_business_day().tomorrow)[:10]
        document.adjust_number_date(self.enNumber, self.enYear, tomorrow)
        # print ('here')
        tableRows = 0

        for oneList in (self.sortedList):
            if oneList:
                tableRows += 2
                tableRows += len(oneList)

        # print (tableRows)
        document.add_new_table(tableRows)
        rowCount = 0
        for num, oneList in enumerate(self.sortedList):
            if oneList:
                tempData = (self.marketsMap[num], '', '', '')
                document.add_table_rows(rowCount, tempData, 'blue')
                rowCount += 1
                document.add_table_rows(rowCount, defaultLayout, 'black')
                rowCount += 1
                n = 0
                for singleLine in oneList:
                    tempData = (singleLine[0],
                                singleLine[1],
                                str(singleLine[2])[:10],
                                singleLine[3])
                    # print (n, str(singleLine[2])[:10])
                    # singleLine[2] = str(singleLine[2])[:10]
                    document.add_table_rows(rowCount, tempData, 'black')
                    rowCount += 1
                    n += 1

        outputName = Super_Duper + 'New_Strikes_' + self.day + '.docx'
        destName = self.monthFolder + 'New_Strikes_' + self.day + '.docx'
        document.save_adjusted(outputName)
        self.convert_docx_pdf(outputName, destName)
        shutil.move(outputName, destName)


    def convert_docx_pdf(self, fileName, destination):

        wordApp = win32.Dispatch('Word.Application')
        wordApp.Visible = False
        destination = destination[:-5] + '.pdf'
        document = wordApp.Documents.Open(fileName)
        fileName = fileName[:-5] + '.pdf'
        document.SaveAs(fileName, FileFormat = 17)
        document.Close()
        # wordApp.Quit()
        shutil.move(fileName, destination)




# test = matchingAlgorithm()


# for num, oneList in enumerate(sortedList):
#     if oneList:
#         print (marketsMap[num])
#         print ('Instrument Series | ISIN Code | Expiration Date | Strike Price')
#         for i in oneList:
#             print (i)