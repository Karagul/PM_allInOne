import win32com.client as win32
import win32timezone
import os
import re
import pytz
from datetime import datetime
from openpyxl import load_workbook
from pandas import to_datetime
from pandas.tseries.offsets import BDay
from bs4 import BeautifulSoup


outlook = win32.Dispatch('Outlook.Application').GetNameSpace('MAPI')

# defInbox = (outlook
#             .Folders('manvydas.kriauciunas@nasdaq.com')
#             .Folders('Inbox')
#             .Folders('PM_DL_test'))
defInbox = outlook.Folders('ListingOperations').Folders('Inbox')

Super_Duper = os.path.expanduser('~\Documents\SUPER DUPER')
Script_Path = os.path.dirname(__file__)

if not os.path.exists(Super_Duper):
    os.makedirs(Super_Duper)


# Take Short Name, add _ to it and paste it to Trading Code column
def add_Trading_Code(xlsxFile):

    sheets = ['Structured Bonds', 'Coupon Bonds']
    wb1 = load_workbook(xlsxFile)
    for sheet in sheets:
        sh1 = wb1.get_sheet_by_name(sheet)
        val = 7
        while True:
            if sh1.cell(row=val, column=1).value != None:                
                name = sh1.cell(row=val, column=1).value[:]
                name = name.replace(' ', '_')

                if sh1.cell(row=6, column=14).value == 'Trading code':
                    sh1.cell(row=val, column=14).value = name
                    date = sh1.cell(row=2, column=8).value
                else:
                    sh1.cell(row=val, column=20).value = name
                    date = sh1.cell(row=2, column=7).value
                val += 1
            else:
                break

        if val > 7:
            break
    wb1.save(xlsxFile)
    

def find_Check_Date(xlsxFile):

    sheets = ['Structured Bonds', 'Coupon Bonds']
    wb1 = load_workbook(xlsxFile, read_only=True)
    for sheet in sheets:
        sh1 = wb1.get_sheet_by_name(sheet)
        if sh1.cell(row=7, column=1).value != None:
            if sh1.cell(row=6, column=14).value == 'Trading code':
                firstDate = sh1.cell(row=2, column=8).value
            else:
                firstDate = sh1.cell(row=2, column=7).value

    strpDate = str(to_datetime(str(firstDate), 
                   format='%Y-%m-%d %H:%M:%S') - 
                   BDay(1))
    checkDate = datetime.strptime((strpDate[:10] + " 13:00"), 
                                           "%Y-%m-%d %H:%M")
    return pytz.utc.localize(checkDate)
    # return firstDate


def get_Instrument_Details(xlsxFile):

    sheets = ['Structured Bonds', 'Coupon Bonds']
    wb1 = load_workbook(xlsxFile, read_only=True)
    for sheet in sheets:
        sh1 = wb1.get_sheet_by_name(sheet)
        val = 7
        while True:
            if sh1.cell(row=val, column=1).value != None:
                val += 1
                if sh1.cell(row=6, column=14).value == 'Trading code':
                    instType = 'SP'
                    instIssuer = str(sh1.cell(row=2, column=3).value)
                    instExchange = (str(sh1.cell(row=2, column=1).value))[1:]
                else:
                    firstDate = sh1.cell(row=2, column=7).value
                    instType = 'COR'
                    instIssuer = str(sh1.cell(row=2, column=3).value)
                    instExchange = (str(sh1.cell(row=2, column=1).value))[1:]
            else:
                break
        if val > 7:
            freeText = (str(val-7) + 
                        ' ' + 
                        instExchange + 
                        ' ' + 
                        instType + 
                        ' ' + 
                        instIssuer)

            return freeText


class Message():

    def __init__(self, message):

        self.message = message

    def replyMessage(self, text, attach=None):

        replyM = self.message.reply
        replyM.GetInspector()

        soup = BeautifulSoup(replyM.HTMLBody, 'lxml')
        tag = soup.find('o:p')

        lines = text.splitlines()

        tag.string = (lines[0])

        for line in lines[1:]:
            tag.append(soup.new_tag('br'))
            tag.append(soup.new_string(line))

        if attach != None:
            replyM.Attachments.Add(attach, 1)

        replyM.HTMLBody = str(soup)

        # replyM.Send()
        replyM.Display(True)

    def __str__(self):
        return self.message.Subject


class BatchingSTO_Message(Message):

    def __init__(self, ID, inbox=defInbox):

        self.ID = ID
        self.xlsxLocation = None
        self.date = None

        messages = inbox.Items

        initMessage = self.findMessage(ID, messages)

        Message.__init__(self, initMessage)

    def findMessage(self, ID, messages):
        
        message = messages.GetLast()

        try:
            while True:
                mesSub = ('Nasdaq Listing Center application '
                          'notification (Record Id {0})')
                if mesSub.format(ID) in message.Subject:
                    return message
                else:
                    message = messages.GetPrevious()
        except:
            pass

    def getRecordID(self):

        return self.ID

    def getAttachments(self):

        attachments = self.message.attachments
        for i in range(1, len(attachments) + 1):
            attachment = attachments.Item(i)
            if (os.path.isfile(Super_Duper + "\\" + 
                               attachment.FileName) == False):
                attachment.SaveAsFile(Super_Duper + "\\" + 
                                      attachment.FileName)
                if attachment.FileName.endswith('.xlsx'):
                    add_Trading_Code(Super_Duper + "\\" + 
                                     attachment.FileName)
                    self.xlsxLocation = (Super_Duper + "\\" + 
                                         attachment.FileName)
            else:
                if attachment.FileName.endswith('.xlsx'):
                    self.xlsxLocation = (Super_Duper + "\\" + 
                                         attachment.FileName)

    def getxlsxLocation(self):
        return self.xlsxLocation

    def getDate(self):
        self.date = find_Check_Date(self.xlsxLocation)
        return self.xlsxLocation

    def createTask(self):

        self.getDate()
        instA_Text = get_Instrument_Details(self.xlsxLocation)

        calendar = outlook.Folders('ListingOperations').Folders('Calendar')

        appointments = calendar.Items

        newApp = appointments.Add(1)
        newApp.Start = self.date
        newApp.Attachments.Add(self.message, 1)
        newApp.Categories = 'Waiting for Exchange notice'
        newApp.Subject = ('List ' + instA_Text + ' ' + self.getRecordID())
        newApp.Body = '/MK'
        # newApp.Body = self.entryInitials.get()
        newApp.Close(0)


class BatchingDK_Message(Message):

    def __init__(self, inbox=defInbox):

        messages = inbox.Items
        self.xmlFileLocations = []

        initMessage = self.getAttachments(messages)

        Message.__init__(self, initMessage)

    def getAttachments(self, messages):

        for message in messages:
            try:
                attachments = message.attachments

                validator = False
                for i in range(1, len(attachments) + 1):

                    attachment = attachments.Item(i)

                    if (attachment.FileName.endswith('.csv') and 
                        'Kladde' in attachment.FileName):
                        if (os.path.isfile(Super_Duper + "\\" + 
                                           attachment.FileName) == False):

                            attachment.SaveAsFile(Super_Duper + "\\" + 
                                                  attachment.FileName)

                            self.xmlFileLocations.append(Super_Duper + "\\" + 
                                                         attachment.FileName)

                            validator = True
                        else:
                            self.xmlFileLocations.append(Super_Duper + "\\" + 
                                                         attachment.FileName)

                            validator = True

                if validator == True:
                    return message
            except:
                print ('error')
                pass

        return None